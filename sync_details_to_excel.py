"""details.js의 상세정보를 작업 폴더의 나우다 엑셀 파일에 반영한다."""

from __future__ import annotations

import copy
import json
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


GALLERY_DIR = Path(__file__).resolve().parent
WORKSPACE_DIR = GALLERY_DIR.parent
DETAIL_HEADERS = [
    "한줄소개",
    "상세소개",
    "운영시간(요일포함)",
    "운영시간 비고",
    "전화번호",
    "상세정보 확인일",
    "상세정보 출처",
    "대표사진 URL",
    "대표사진 대체텍스트",
    "대표사진 출처",
]

STATUS_NOTES = {
    "라반 이호테우점": "확인 필요: 동일 주소에 다른 상호의 계속사업자 정보가 확인되어 라반의 폐업·상호 변경 여부 확인 필요 (2026-07-14)",
    "디저트팩토리 쇠소깍점": "휴업/폐업: 비짓제주 나우다 공식 API에 '업체 폐업으로 이용 불가' 명시 (2026-07-14)",
    "제주지프": "확인 필요: 비짓제주 나우다 공식 API에 '동절기 기간 임시 휴업' 문구가 현재도 남아 있어 운영 재개 여부 확인 필요 (2026-07-14)",
}


def load_details() -> dict[int, dict]:
    text = (GALLERY_DIR / "details.js").read_text(encoding="utf-8")
    body = text.split("=", 1)[1].strip().removesuffix(";")
    try:
        parsed = json.loads(body)
    except json.JSONDecodeError:
        body = re.sub(
            r"([,{]\s*)([A-Za-z_][A-Za-z0-9_]*|\d+)\s*:",
            r'\1"\2":',
            body,
        )
        parsed = json.loads(body)
    return {int(key): value for key, value in parsed.items()}


def load_images() -> dict[int, dict]:
    text = (GALLERY_DIR / "images.js").read_text(encoding="utf-8")
    body = text.split("=", 1)[1].strip().removesuffix(";")
    return {int(key): value for key, value in json.loads(body).items()}


def load_data() -> list[dict]:
    text = (GALLERY_DIR / "data.js").read_text(encoding="utf-8")
    return json.loads(text.split("=", 1)[1].strip().removesuffix(";"))


def hours_text(detail: dict) -> str:
    return "\n".join(
        f'{row["days"]}  {row["hours"]}' for row in detail.get("weeklyHours", [])
    )


def update_workbook(
    path: Path,
    details_by_name: dict[str, dict],
    data_by_name: dict[str, dict],
) -> None:
    workbook = load_workbook(path)
    sheet = workbook.active
    expected_rows = len(data_by_name) + 1
    if sheet.max_row not in (243, expected_rows):
        raise ValueError(f"{path.name}: 예상하지 못한 제휴사 행 수입니다 ({sheet.max_row - 1}개)")

    headers = {sheet.cell(1, column).value: column for column in range(1, sheet.max_column + 1)}
    header_template = sheet.cell(1, sheet.max_column)
    for header in DETAIL_HEADERS:
        if header not in headers:
            column = sheet.max_column + 1
            cell = sheet.cell(1, column, header)
            cell._style = copy.copy(header_template._style)
            cell.font = copy.copy(header_template.font)
            cell.fill = copy.copy(header_template.fill)
            cell.border = copy.copy(header_template.border)
            cell.alignment = copy.copy(header_template.alignment)
            cell.number_format = header_template.number_format
            headers[header] = column

    name_column = headers["업체명"]
    address_column = headers["주소"]
    updated_names: set[str] = set()

    for row in range(2, sheet.max_row + 1):
        name = sheet.cell(row, name_column).value
        detail = details_by_name.get(name)
        if not detail:
            continue

        updated_names.add(name)
        if detail.get("address"):
            sheet.cell(row, address_column, detail["address"])

        values = {
            "한줄소개": detail.get("summary", ""),
            "상세소개": detail.get("description", ""),
            "운영시간(요일포함)": hours_text(detail),
            "운영시간 비고": detail.get("hoursNote", ""),
            "전화번호": detail.get("phone", ""),
            "상세정보 확인일": detail.get("checkedAt", ""),
            "상세정보 출처": "\n".join(
                value
                for value in (detail.get("infoSourceLabel"), detail.get("infoSourceUrl"))
                if value
            ),
            "대표사진 URL": detail.get("imageUrl", ""),
            "대표사진 대체텍스트": detail.get("imageAlt", ""),
            "대표사진 출처": "\n".join(
                value
                for value in (detail.get("imageSourceLabel"), detail.get("imageSourceUrl"))
                if value
            ),
        }
        for header, value in values.items():
            cell = sheet.cell(row, headers[header], value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        if name == "제주관광공사 중문면세점" and "비고" in headers:
            sheet.cell(
                row,
                headers["비고"],
                "주소 표기 정정: 공식 안내 주소 '서귀포시 중문관광로 224, ICC JEJU 1층' 반영 (2026-07-13)",
            ).alignment = Alignment(vertical="top", wrap_text=True)
        if name in STATUS_NOTES and "비고" in headers:
            sheet.cell(
                row,
                headers["비고"],
                STATUS_NOTES[name],
            ).alignment = Alignment(vertical="top", wrap_text=True)

    missing = set(details_by_name) - updated_names
    if missing:
        template_row = sheet.max_row
        for name in sorted(missing, key=lambda value: data_by_name[value]["id"]):
            item = data_by_name[name]
            detail = details_by_name[name]
            row = sheet.max_row + 1
            for column in range(1, sheet.max_column + 1):
                source = sheet.cell(template_row, column)
                target = sheet.cell(row, column)
                target._style = copy.copy(source._style)
                target.font = copy.copy(source.font)
                target.fill = copy.copy(source.fill)
                target.border = copy.copy(source.border)
                target.alignment = copy.copy(source.alignment)
                target.number_format = source.number_format
            core_values = {
                "업체명": name,
                "카테고리": item.get("category", ""),
                "주소": detail.get("address") or item.get("address", ""),
                "혜택": item.get("benefit", ""),
                "상세링크": item.get("sourceUrl", ""),
                "한줄소개": detail.get("summary", ""),
                "상세소개": detail.get("description", ""),
                "운영시간(요일포함)": hours_text(detail),
                "운영시간 비고": detail.get("hoursNote", ""),
                "전화번호": detail.get("phone", ""),
                "상세정보 확인일": detail.get("checkedAt", ""),
                "상세정보 출처": "\n".join(
                    value for value in (detail.get("infoSourceLabel"), detail.get("infoSourceUrl")) if value
                ),
                "대표사진 URL": detail.get("imageUrl", ""),
                "대표사진 대체텍스트": detail.get("imageAlt", ""),
                "대표사진 출처": "\n".join(
                    value for value in (detail.get("imageSourceLabel"), detail.get("imageSourceUrl")) if value
                ),
            }
            for header, value in core_values.items():
                if header in headers:
                    sheet.cell(row, headers[header], value).alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        "한줄소개": 38,
        "상세소개": 62,
        "운영시간(요일포함)": 31,
        "운영시간 비고": 52,
        "전화번호": 18,
        "상세정보 확인일": 17,
        "상세정보 출처": 47,
        "대표사진 URL": 65,
        "대표사진 대체텍스트": 34,
        "대표사진 출처": 55,
    }
    for header, width in widths.items():
        sheet.column_dimensions[get_column_letter(headers[header])].width = width
    if sheet.auto_filter.ref:
        sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"

    temporary = path.with_name(f".{path.stem}.updating{path.suffix}")
    workbook.save(temporary)
    check = load_workbook(temporary, read_only=True, data_only=True)
    check_sheet = check.active
    check_headers = [check_sheet.cell(1, column).value for column in range(1, check_sheet.max_column + 1)]
    if check_sheet.max_row != expected_rows or not all(header in check_headers for header in DETAIL_HEADERS):
        check.close()
        temporary.unlink(missing_ok=True)
        raise ValueError(f"{path.name}: 저장 후 검증에 실패했습니다")
    check.close()
    temporary.replace(path)


def main() -> None:
    details = load_details()
    images = load_images()
    data = load_data()
    names = {item["id"]: item["name"] for item in data}
    data_by_name = {item["name"]: item for item in data}
    records = {
        item_id: {**image, **details.get(item_id, {})}
        for item_id, image in images.items()
    }
    details_by_name = {names[item_id]: detail for item_id, detail in records.items()}
    targets = sorted(WORKSPACE_DIR.glob("나우다_관광사업체_242개*.xlsx"))
    if not targets:
        raise FileNotFoundError("업데이트할 나우다 엑셀 파일이 없습니다")
    for target in targets:
        update_workbook(target, details_by_name, data_by_name)
        print(f"UPDATED {target.name}: {len(details_by_name)}개 상세정보")


if __name__ == "__main__":
    main()
